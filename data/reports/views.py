import io
from datetime import datetime

from django.http import HttpResponse
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from apps.common.auth.authentication import UnifiedJWTAuthentication
from apps.common.permissions import HasDynamicPermission
from data.orders.models import Order
from data.excavator.models import ExcavatorOrder
from data.products.models import WhouseProducts


COMPANY = '«Қодир Инвест Сервис» МЧЖ'
PHONES = '(+998) 94 444 05 38 Сейитжан   (+998) 77 707 71 72 Бегзат'

STATUS_UZ = {
    'NEW': 'Янги',
    'IN_PROGRESS': 'Жараёнда',
    'ON_WAY': "Йўлда",
    'ARRIVED': 'Етиб келди',
    'UNLOADING': 'Туширилмоқда',
    'COMPLETED': 'Етказиб берилган',
    'REJECTED': 'Бекор қилинган',
    'PAUSED': "Тўхтатилган",
    'EXPIRED': 'Муддати тугаган',
}

PAYMENT_STATUS_UZ = {
    'PENDING': "Тўланмаган",
    'PAID': "Тўланган",
}


def status_uz(s):
    return STATUS_UZ.get(s, s)


def fmt_date(dt):
    if not dt:
        return ''
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
        except Exception:
            return str(dt)
    try:
        return dt.strftime('%d.%m.%Y')
    except Exception:
        return str(dt)


def get_history_time(history, stat):
    if not history:
        return ''
    for entry in history:
        if entry.get('status') == stat:
            ts = entry.get('timestamp')
            if ts:
                try:
                    if isinstance(ts, str):
                        dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
                    else:
                        dt = ts
                    return dt.strftime('%H:%M')
                except Exception:
                    return str(ts)
    return ''


def px(pixels):
    return round(pixels / 7.0, 1)


_thin = Side(style='thin', color='000000')
_none = Side(style=None)
ALL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _font(bold=False, size=8):
    return Font(name='Calibri', size=size, bold=bold)


def _align(h='center', wrap=False):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)


def style_range(ws, min_row, min_col, max_row, max_col, bold=False, h='center', wrap=False, border=False):
    f = _font(bold)
    a = _align(h, wrap)
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = f
            cell.alignment = a
            if border:
                cell.border = ALL_BORDER


def outer_border(ws, min_row, min_col, max_row, max_col, top=True, left=True, bottom=True, right=True):
    rows = list(ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col))
    n_rows = len(rows)
    for ri, row in enumerate(rows):
        n_cols = len(row)
        for ci, cell in enumerate(row):
            t = _thin if (top and ri == 0) else _none
            b = _thin if (bottom and ri == n_rows - 1) else _none
            l = _thin if (left and ci == 0) else _none
            r = _thin if (right and ci == n_cols - 1) else _none
            cell.border = Border(top=t, bottom=b, left=l, right=r)


def apply_outer_frame(ws, min_row, min_col, max_row, max_col):
    """Mavjud borderlarga tegmasdan faqat tashqi ramkani qo'shadi."""
    rows = list(ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col))
    n_rows = len(rows)
    for ri, row in enumerate(rows):
        n_cols = len(row)
        for ci, cell in enumerate(row):
            e = cell.border
            cell.border = Border(
                top=_thin if ri == 0 else e.top,
                bottom=_thin if ri == n_rows - 1 else e.bottom,
                left=_thin if ci == 0 else e.left,
                right=_thin if ci == n_cols - 1 else e.right,
            )


def merge_val(ws, range_str, val, bold=False, h='center', wrap=False,
              top=True, left=True, bottom=True, right=True):
    ws.merge_cells(range_str)
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    cell = ws.cell(row=min_row, column=min_col)
    cell.value = val
    cell.font = _font(bold)
    cell.alignment = _align(h, wrap)
    outer_border(ws, min_row, min_col, max_row, max_col, top, left, bottom, right)
    return cell


def make_excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Sheet 1: Yuk xati ───────────────────────────────────────────────────────

def fill_yuk_xati(ws, order):
    ws.title = 'Юк хати'

    for i, w in enumerate([20, 72, 46, 90, 72, 85], 1):
        ws.column_dimensions[get_column_letter(i)].width = px(w)

    sub_orders = list(order.sub_orders.all())

    merge_val(ws, 'A1:G1', 'ЮК ХАТИ', bold=True, bottom=False)
    merge_val(ws, 'A2:G2', f'Кимдан:  {COMPANY}', top=False, bottom=False)
    merge_val(ws, 'A3:G3', f'Кимга: {order.client.name}  ИНН: {order.client.inn_number}',
              top=False)

    # Row 4 headers
    ws.merge_cells('B4:C4')
    for col, val in [(1, '№'), (2, 'Махсулот номи'), (4, 'Ўлчов бирлиги'),
                     (5, 'Микдори'), (6, 'Нархи'), (7, 'Суммаси')]:
        ws.cell(row=4, column=col, value=val)
    style_range(ws, 4, 1, 4, 7, border=True)

    # Rows 5+: order_items data
    order_items = list(order.order_items.all())
    for idx, oi in enumerate(order_items):
        row = 5 + idx
        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row=row, column=1, value=idx + 1)
        ws.cell(row=row, column=2, value=f"{oi.product.name} {oi.type.name}")
        ws.cell(row=row, column=4, value=oi.unit.name)
        ws.cell(row=row, column=5, value=float(oi.quantity))
        price_cell = ws.cell(row=row, column=6, value=float(oi.price))
        price_cell.number_format = '#,##0'
        total_cell = ws.cell(row=row, column=7, value=float(oi.quantity) * float(oi.price))
        total_cell.number_format = '#,##0'
        style_range(ws, row, 1, row, 7, border=True)

    items_end = 5 + max(len(order_items), 1) - 1

    # Yetkazib beruvchilar title
    yb_title_row = items_end + 1
    merge_val(ws, f'A{yb_title_row}:G{yb_title_row}', 'Етказиб берувчилар', bold=True, top=False, bottom=False)

    # sub_orders table headers
    yb_head_row = yb_title_row + 1
    ws.row_dimensions[yb_head_row].height = 36
    yb_headers = [
        '№', 'автомобил рақами', 'шафёр', 'юк миқдори (куб)',
        'заводдан \nчиқарилган вақти', 'Маҳсулот етказиб \nберилган вақти', 'Юк тушириб \nолинган вақти',
    ]
    for ci, val in enumerate(yb_headers, 1):
        c = ws.cell(row=yb_head_row, column=ci, value=val)
        c.font = _font()
        c.alignment = _align(wrap=True)
        c.border = ALL_BORDER

    # sub_order data rows
    yb_data_start = yb_head_row + 1
    for i, so in enumerate(sub_orders):
        row = yb_data_start + i
        history = so.status_history or []
        so_qty = sum(float(si.quantity) for si in so.sub_order_items.all())
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=so.transport.number if so.transport else '')
        ws.cell(row=row, column=3, value=so.driver.name if so.driver else '')
        ws.cell(row=row, column=4, value=so_qty)
        ws.cell(row=row, column=5, value=get_history_time(history, 'ON_WAY'))
        ws.cell(row=row, column=6, value=get_history_time(history, 'ARRIVED'))
        ws.cell(row=row, column=7, value=get_history_time(history, 'COMPLETED'))
        style_range(ws, row, 1, row, 7, border=True)

    # Qabul qiluvchilar section
    qb_title_row = yb_data_start + len(sub_orders)
    qb_head_row = qb_title_row + 1
    qb_data_start = qb_head_row + 1

    merge_val(ws, f'A{qb_title_row}:F{qb_title_row}', 'Қабул қилувчилар',
              bold=True, top=False, bottom=False)

    qb_headers = [
        '№', 'автомобил', 'шафёр',
        'Қабул қилувчи (исм)', 'Қабул қилувчи (лавозими)', 'ҳолати',
    ]
    for ci, val in enumerate(qb_headers, 1):
        c = ws.cell(row=qb_head_row, column=ci, value=val)
        c.font = _font()
        c.alignment = _align()
        c.border = ALL_BORDER

    for i, so in enumerate(sub_orders):
        row = qb_data_start + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=so.transport.number if so.transport else '')
        ws.cell(row=row, column=3, value=so.driver.name if so.driver else '')
        ws.cell(row=row, column=4, value=order.client.name)
        ws.cell(row=row, column=5, value='мижоз')
        ws.cell(row=row, column=6, value='Қабул қилдим' if so.status == 'COMPLETED' else status_uz(so.status))
        style_range(ws, row, 1, row, 6, border=True)

    # Bottom
    bottom_row = qb_data_start + len(sub_orders)
    branch_addr = order.branch.address if order.branch else ''
    ws.merge_cells(f'B{bottom_row}:G{bottom_row}')
    addr_cell = ws.cell(row=bottom_row, column=2, value=f'Манзил: {branch_addr}')
    addr_cell.font = _font()
    addr_cell.alignment = _align('left', wrap=True)
    ws.merge_cells(f'B{bottom_row + 1}:G{bottom_row + 1}')
    phone_cell = ws.cell(row=bottom_row + 1, column=2, value=PHONES)
    phone_cell.font = _font()
    phone_cell.alignment = _align('left')
    imzo_cell = ws.cell(row=bottom_row + 2, column=5, value='Имзо:')
    imzo_cell.font = _font()
    imzo_cell.alignment = _align()

    apply_outer_frame(ws, 1, 1, bottom_row + 2, 7)


# ─── Sheet 2: Ishonch qog'ozi ────────────────────────────────────────────────

def fill_ishonch_qogozi(ws, order, sub_order, idx=None):
    ws.title = "Ишонч қоғози" if idx is None else f"Ишонч қоғози {idx}"

    for i, w in enumerate([28, 156, 76, 51, 61, 98], 1):
        ws.column_dimensions[get_column_letter(i)].width = px(w)

    merge_val(ws, 'A1:F1', 'ЮК ХАТИ', bold=True, bottom=False)
    merge_val(ws, 'A2:F2', f'Кимдан:  {COMPANY}', h='left', top=False, bottom=False)

    date_cell = ws.cell(row=3, column=6, value=fmt_date(order.created_at))
    date_cell.font = _font()
    date_cell.alignment = _align()
    outer_border(ws, 3, 6, 3, 6, top=False, left=False, bottom=False, right=True)

    merge_val(ws, 'A4:C4', f'Кимга: {order.client.name}',
              h='left', top=False, bottom=False, right=False)
    merge_val(ws, 'D4:F4', f'ИНН: {order.client.inn_number}',
              h='left', top=False, left=False, bottom=False)

    driver_name = sub_order.driver.name if sub_order.driver else ''
    transport_num = sub_order.transport.number if sub_order.transport else ''
    merge_val(ws, 'A5:F5', f"Ишонч қоғози:  {driver_name} / {transport_num}",
              h='left', top=False, bottom=False)

    # Row 6 headers
    for ci, val in enumerate(['№', 'Маҳсулот номи', "Ўлчов бирлиги", 'Миқдори', 'нархи', 'Суммаси'], 1):
        c = ws.cell(row=6, column=ci, value=val)
        c.font = _font()
        c.alignment = _align()
        c.border = ALL_BORDER

    # Rows 7+: sub_order_items data
    sub_items = list(sub_order.sub_order_items.all())
    # Build a price lookup from order items: (product_id, type_id, unit_id) -> price
    price_map = {
        (oi.product_id, oi.type_id, oi.unit_id): oi.price
        for oi in order.order_items.all()
    }
    for si_idx, si in enumerate(sub_items):
        row = 7 + si_idx
        price = float(price_map.get((si.product_id, si.type_id, si.unit_id), 0))
        ws.cell(row=row, column=1, value=si_idx + 1)
        ws.cell(row=row, column=2, value=f"{si.product.name} {si.type.name}")
        ws.cell(row=row, column=3, value=si.unit.name)
        ws.cell(row=row, column=4, value=float(si.quantity))
        price_cell = ws.cell(row=row, column=5, value=price)
        price_cell.number_format = '#,##0'
        total_cell = ws.cell(row=row, column=6, value=float(si.quantity) * price)
        total_cell.number_format = '#,##0'
        style_range(ws, row, 1, row, 6, border=True)

    items_last_row = 7 + max(len(sub_items), 1) - 1

    # Empty signature rows
    sig_start = items_last_row + 1
    for r in [sig_start, sig_start + 1]:
        ws.cell(row=r, column=1, value=r - sig_start + len(sub_items) + 1)
        style_range(ws, r, 1, r, 6, border=True)

    phones_row = sig_start + 2
    merge_val(ws, f'A{phones_row}:F{phones_row}', PHONES, h='left', top=False, bottom=False)

    bottom_row = phones_row + 1
    merge_val(ws, f'A{bottom_row}:B{bottom_row}', 'Топширди:', h='left', top=False, right=False)
    merge_val(ws, f'C{bottom_row}:F{bottom_row}', 'Қабул қилди:', h='left', top=False, left=False)


# ─── Sheet 3: Buyurtmalar hisoboti ───────────────────────────────────────────

def fill_buyurtmalar_hisoboti(ws, orders):
    ws.title = "Буюртмалар бўйича ҳисобот"

    for i, w in enumerate([27, 153, 69, 40, 30, 40, 79, 90, 142], 1):
        ws.column_dimensions[get_column_letter(i)].width = px(w)

    ws.merge_cells('A1:I1')
    ws['A1'].value = f"Буюртмалар бўйича ҳисобот {COMPANY}"
    ws['A1'].font = Font(name='Calibri', size=8)
    ws['A1'].alignment = _align()

    for ci, val in enumerate(['№', 'Мижозлар', 'Маҳсулот номи', 'Миқдори', 'Бирлиги',
                               'Нархи', 'Суммаси', 'Сана', 'Ҳолати'], 1):
        c = ws.cell(row=2, column=ci, value=val)
        c.font = _font()
        c.alignment = _align()
        c.border = ALL_BORDER

    row_num = 3
    grand_total = 0
    for order in orders:
        for oi in order.order_items.all():
            type_name = oi.type.name if oi.type else ''
            ws.cell(row=row_num, column=1, value=row_num - 2)
            ws.cell(row=row_num, column=2, value=order.client.name)
            ws.cell(row=row_num, column=3, value=f"{oi.product.name} {type_name}".strip())
            ws.cell(row=row_num, column=4, value=float(oi.quantity))
            ws.cell(row=row_num, column=5, value=oi.unit.name if oi.unit else '')
            price_c = ws.cell(row=row_num, column=6, value=float(oi.price))
            price_c.number_format = '#,##0'
            line_total = float(oi.quantity) * float(oi.price)
            total_c = ws.cell(row=row_num, column=7, value=line_total)
            total_c.number_format = '#,##0'
            ws.cell(row=row_num, column=8, value=fmt_date(order.created_at))
            ws.cell(row=row_num, column=9, value=status_uz(order.status))
            style_range(ws, row_num, 1, row_num, 9, border=True)
            grand_total += line_total
            row_num += 1

    if row_num > 3:
        ws.merge_cells(f'A{row_num}:F{row_num}')
        jami_cell = ws.cell(row=row_num, column=1, value='ЖАМИ:')
        jami_cell.font = _font(bold=True)
        jami_cell.alignment = _align('right')
        gtotal_c = ws.cell(row=row_num, column=7, value=grand_total)
        gtotal_c.number_format = '#,##0'
        gtotal_c.font = _font(bold=True)
        style_range(ws, row_num, 1, row_num, 9, border=True)


# ─── Sheet 4: Yetkazib beruvchilar hisoboti ──────────────────────────────────

def fill_yetkazib_beruvchilar_hisoboti(ws, items):
    ws.title = "Етказиб берувчилар бўйича ҳисобот"

    for i, w in enumerate([27, 161, 99, 119, 74, 81, 77, 94], 1):
        ws.column_dimensions[get_column_letter(i)].width = px(w)

    ws.merge_cells('A1:H1')
    ws['A1'].value = "Етказиб берувчилар бўйича ҳисобот"
    ws['A1'].font = Font(name='Calibri', size=8)
    ws['A1'].alignment = _align()

    for ci, val in enumerate(['№', 'Етказиб берувчи', 'Номери', 'Маҳсулот номи',
                               'Тури', 'Миқдори', 'Бирлиги', 'Сана'], 1):
        c = ws.cell(row=2, column=ci, value=val)
        c.font = _font()
        c.alignment = _align()
        c.border = ALL_BORDER

    for i, item in enumerate(items):
        row = i + 3
        supplier = item.supplier
        phone = ''
        if supplier:
            first_phone = supplier.phones.first()
            if first_phone:
                phone = first_phone.phone_number
        unit_name = item.product.unit.name if item.product and item.product.unit else ''
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=supplier.name if supplier else '')
        ws.cell(row=row, column=3, value=phone)
        ws.cell(row=row, column=4, value=item.product.name if item.product else '')
        ws.cell(row=row, column=5, value=item.product_type.name if item.product_type else '-')
        ws.cell(row=row, column=6, value=float(item.quantity))
        ws.cell(row=row, column=7, value=unit_name)
        ws.cell(row=row, column=8, value=fmt_date(item.created_at))
        style_range(ws, row, 1, row, 8, border=True)


# ─── API Views ────────────────────────────────────────────────────────────────

class YukXatiExcelView(APIView):
    authentication_classes = [UnifiedJWTAuthentication]
    permission_classes = [HasDynamicPermission(read_perm="ORDERS_PAGE")]

    def get(self, request, pk):
        order = (
            Order.objects
            .select_related('client', 'branch')
            .prefetch_related(
                'order_items__product', 'order_items__type', 'order_items__unit',
                'sub_orders__driver', 'sub_orders__transport',
                'sub_orders__sub_order_items__product',
                'sub_orders__sub_order_items__type',
                'sub_orders__sub_order_items__unit',
            )
            .filter(id=pk)
            .first()
        )
        if not order:
            return HttpResponse(status=404)
        wb = Workbook()
        fill_yuk_xati(wb.active, order)
        return make_excel_response(wb, f'юк_хати_{order.display_id}.xlsx')


class IshonchQogoziExcelView(APIView):
    authentication_classes = [UnifiedJWTAuthentication]
    permission_classes = [HasDynamicPermission(read_perm="ORDERS_PAGE")]

    def get(self, request, pk):
        order = (
            Order.objects
            .select_related('client', 'branch')
            .prefetch_related(
                'order_items__product', 'order_items__type', 'order_items__unit',
                'sub_orders__driver', 'sub_orders__transport',
                'sub_orders__sub_order_items__product',
                'sub_orders__sub_order_items__type',
                'sub_orders__sub_order_items__unit',
            )
            .filter(id=pk)
            .first()
        )
        if not order:
            return HttpResponse(status=404)
        sub_orders = list(order.sub_orders.all())
        wb = Workbook()
        for i, so in enumerate(sub_orders):
            ws = wb.active if i == 0 else wb.create_sheet()
            idx = None if len(sub_orders) == 1 else i + 1
            fill_ishonch_qogozi(ws, order, so, idx)
        return make_excel_response(wb, f'ишонч_қоғози_{order.display_id}.xlsx')


class BuyurtmalarHisobotiExcelView(APIView):
    authentication_classes = [UnifiedJWTAuthentication]
    permission_classes = [HasDynamicPermission(read_perm="ORDERS_PAGE")]

    def get(self, request):
        qs = Order.objects.select_related('client').prefetch_related(
            'order_items__product', 'order_items__type', 'order_items__unit',
        )
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)
        wb = Workbook()
        fill_buyurtmalar_hisoboti(wb.active, list(qs))
        return make_excel_response(wb, 'буюртмалар_ҳисоботи.xlsx')


class YetkazibBeruvchilarHisobotiExcelView(APIView):
    authentication_classes = [UnifiedJWTAuthentication]
    permission_classes = [HasDynamicPermission(read_perm="ORDERS_PAGE")]

    def get(self, request):
        qs = WhouseProducts.objects.select_related(
            'supplier', 'product__unit', 'product_type'
        ).prefetch_related('supplier__phones')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)
        wb = Workbook()
        fill_yetkazib_beruvchilar_hisoboti(wb.active, list(qs))
        return make_excel_response(wb, 'етказиб_берувчилар_ҳисоботи.xlsx')


# ─── Sheet 5: Ekskavator buyurtmalari hisoboti ──────────────────────────────

def fill_excavator_hisoboti(ws, orders):
    ws.title = "Экскаватор буюртмалари"

    for i, w in enumerate([27, 120, 80, 70, 70, 70, 80, 80, 100, 80], 1):
        ws.column_dimensions[get_column_letter(i)].width = px(w)

    ws.merge_cells('A1:J1')
    ws['A1'].value = f"Экскаватор буюртмалари бўйича ҳисобот {COMPANY}"
    ws['A1'].font = Font(name='Calibri', size=8)
    ws['A1'].alignment = _align()

    headers = [
        '№', 'Мижоз', 'Телефон', 'Манзил',
        'Бошланиш', 'Тугаш', 'Ҳолати',
        "Тўлов", 'Ҳайдовчи', 'Транспорт',
    ]
    for ci, val in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=val)
        c.font = _font()
        c.alignment = _align()
        c.border = ALL_BORDER

    row_num = 3
    for order in orders:
        sub_orders = list(order.sub_orders.all())
        if not sub_orders:
            ws.cell(row=row_num, column=1, value=row_num - 2)
            ws.cell(row=row_num, column=2, value=order.client_name)
            ws.cell(row=row_num, column=3, value=order.phone_number)
            ws.cell(row=row_num, column=4, value=order.address or '')
            ws.cell(row=row_num, column=5, value=fmt_date(order.start_date))
            ws.cell(row=row_num, column=6, value=fmt_date(order.end_date))
            ws.cell(row=row_num, column=7, value=status_uz(order.status))
            ws.cell(row=row_num, column=8, value=PAYMENT_STATUS_UZ.get(order.payment_status, order.payment_status))
            ws.cell(row=row_num, column=9, value='')
            ws.cell(row=row_num, column=10, value='')
            style_range(ws, row_num, 1, row_num, 10, border=True)
            row_num += 1
        else:
            for so in sub_orders:
                ws.cell(row=row_num, column=1, value=row_num - 2)
                ws.cell(row=row_num, column=2, value=order.client_name)
                ws.cell(row=row_num, column=3, value=order.phone_number)
                ws.cell(row=row_num, column=4, value=order.address or '')
                ws.cell(row=row_num, column=5, value=fmt_date(order.start_date))
                ws.cell(row=row_num, column=6, value=fmt_date(order.end_date))
                ws.cell(row=row_num, column=7, value=status_uz(order.status))
                ws.cell(row=row_num, column=8, value=PAYMENT_STATUS_UZ.get(order.payment_status, order.payment_status))
                ws.cell(row=row_num, column=9, value=so.driver.name if so.driver else '')
                ws.cell(row=row_num, column=10, value=so.transport.number if so.transport else '')
                style_range(ws, row_num, 1, row_num, 10, border=True)
                row_num += 1

    if row_num > 3:
        ws.merge_cells(f'A{row_num}:H{row_num}')
        jami_cell = ws.cell(row=row_num, column=1, value=f'ЖАМИ: {row_num - 3} та')
        jami_cell.font = _font(bold=True)
        jami_cell.alignment = _align('right')
        style_range(ws, row_num, 1, row_num, 10, border=True)


class ExcavatorHisobotiExcelView(APIView):

    def get(self, request):
        qs = ExcavatorOrder.objects.prefetch_related(
            'sub_orders__driver', 'sub_orders__transport',
        )
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)
        wb = Workbook()
        fill_excavator_hisoboti(wb.active, list(qs))
        return make_excel_response(wb, 'экскаватор_ҳисоботи.xlsx')
